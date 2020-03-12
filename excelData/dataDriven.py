class xlData:
    import openpyxl
    def readData(self,sheetname,row,col,filepath):
        obj=self.openpyxl.load_workbook(filepath)
        s=obj.get_sheet_by_name(sheetname)
        res=s.cell(row=row,column=col).value
        return res

    def writeData(self,sheetname,row,col,filepath,data):
        obj=self.openpyxl.load_workbook(filepath)
        s=obj.get_sheet_by_name(sheetname)
        s.cell(row=row,column=col).value=data

    def maxRowCount(self):
        #from openpyxl import Workbook
        #book = Workbook()
        #sheet = book.active
        sheet=self.openpyxl.Workbook().active
        return sheet.max_row

    def maxColCount(self):
        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active
        return sheet.max_column

