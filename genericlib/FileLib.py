import openpyxl
import json
from openpyxl import Workbook

class fileData:
    filepath_User = r'C:\Users\Vidyashree\PycharmProjects\makeMyTrip\username_pwd.xlsx'
    filepath_flightsheet = r'C:\Users\Vidyashree\PycharmProjects\makeMyTrip\testData\flight_sheet.xlsx'

    def readxlData(self,sheetname,row,col,filepath):
        obj=openpyxl.load_workbook(filepath)
        s=obj.get_sheet_by_name(sheetname)
        res=s.cell(row=row,column=col).value
        return res

    def writexlData(self,sheetname,row,col,filepath,data):
        obj=openpyxl.load_workbook(filepath)
        s=obj.get_sheet_by_name(sheetname)
        s.cell(row=row,column=col).value=data
        obj.save(filepath)

    def xlmaxRowCount(self,sheetname):
        sheet=openpyxl.Workbook().get_sheet_by_name(sheetname)
        return sheet.max_row

    def xlmaxColCount(self ,sheetname):
        book = Workbook()
        sheet = book.get_sheet_by_name(sheetname)
        return sheet.max_column

    def readJson(self,jsonPath):
        with open(jsonPath, 'r') as file:
            credentials = json.load(file)
        return credentials

    def writeJson(self,jsonPath,data):
        with open(jsonPath,'w') as file:
            Data=json.dump(data,file)
        return Data








































    # def xlReadData(self,sheetname,file):
    #     objfile = fileData()
    #     maxRow = objfile.maxRowCount()
    #     maxCol=objfile.maxColCount()
    #     data=[]
    #
    #     for i in range(1, maxRow + 1):
    #         for j in range(1,maxCol+1):
    #             data.append(objfile.readData(sheetname, i + 1, j, file))
    #
    #         return data
    #
    #
    # def user_pwdData(self):
    #     maxRow = self.maxRowCount()
    #     for i in range(1, maxRow + 1):
    #         username = self.readData('Sheet1', i + 1, 1, self.filepath_User)
    #         password = self.readData('Sheet1', i + 1, 2,self.filepath_User )
    #         return username,password
